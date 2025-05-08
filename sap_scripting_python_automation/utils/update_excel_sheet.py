import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
import datetime
import logging

logger = logging.getLogger(__name__)


def update_excel_with_sap_data(df: pd.DataFrame, EXCEL_FILE_PATH: str, SHEET_NAME: str,
                               date_columns: list = None, none_df: pd.DataFrame = None) -> bool:
    """
    Update the specified sheet in the Excel file with the provided DataFrame.
    If the DataFrame is empty (only column names), it will still update the sheet.
    If the DataFrame is None, it will use the provided `none_df` to update the sheet.

    For each column specified in the date_columns list, the column is converted to a date,
    and its cells are displayed in the MM/DD/YYYY format in Excel.

    If a cell in a date column already has the MM/DD/YYYY style applied, it will not be modified.
    """
    try:
        # Handle None input by replacing it with none_df
        if df is None:
            if none_df is not None:
                df = none_df
                logger.info("Input DataFrame is None. Using the provided `none_df` to update the Excel sheet.")
            else:
                logger.warning("Input DataFrame is None and no `none_df` provided. Skipping Excel update.")
                return False

        logger.info(f"Updating Excel file: {EXCEL_FILE_PATH}")

        # Convert numeric columns to float (if applicable) and ignore conversion errors.
        for column in df.select_dtypes(include=['object']).columns:
            df[column] = pd.to_numeric(df[column], errors='ignore')

        # Default to an empty list if None is provided
        date_columns = date_columns or []

        # Convert date columns
        for col in date_columns:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce').dt.date

        # Preserve the original column order.
        df = df.reindex(columns=df.columns)

        # Load the workbook.
        wb = load_workbook(EXCEL_FILE_PATH)

        # Check if the sheet exists; if not, create it.
        if SHEET_NAME not in wb.sheetnames:
            logger.warning(f"Sheet '{SHEET_NAME}' not found in Excel. Creating new sheet.")
            ws = wb.create_sheet(SHEET_NAME)
        else:
            ws = wb[SHEET_NAME]
            ws.delete_rows(1, ws.max_row)  # Clear existing data.

        # Write headers.
        ws.append(df.columns.tolist())

        # Write the DataFrame rows.
        for row in df.itertuples(index=False, name=None):
            ws.append(row)

        # -------------------------------
        #  Apply MM/DD/YYYY Format
        # -------------------------------
        # Create a NamedStyle for MM/DD/YYYY if not already present.
        date_style = NamedStyle(name="mmddyyyy")
        date_style.number_format = "MM/DD/YYYY"

        # Safely collect existing named style names
        existing_named_styles = [style.name for style in wb.named_styles if isinstance(style, NamedStyle)]
        if date_style.name not in existing_named_styles:
            wb.named_styles.append(date_style)

        # Loop through each provided date column and apply the date style.
        for col in date_columns:
            if col in df.columns:
                date_col_idx = df.columns.get_loc(col) + 1  # +1 because openpyxl columns start at 1
                for row_idx in range(2, df.shape[0] + 2):
                    cell = ws.cell(row=row_idx, column=date_col_idx)
                    # Apply style only if the cell contains a date and is not already styled.
                    if isinstance(cell.value, datetime.date) and cell.style != date_style.name:
                        cell.style = date_style

        # Save the updated workbook.
        wb.save(EXCEL_FILE_PATH)
        logger.info(f"Successfully updated '{SHEET_NAME}' in {EXCEL_FILE_PATH}")
        return True

    except Exception as e:
        logger.error(f"Error updating Excel file: {e}", exc_info=True)
        return False
